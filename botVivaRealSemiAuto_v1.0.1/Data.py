import os
import sys

import urllib3
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, NoSuchElementException

from bs4 import BeautifulSoup

from time import sleep
from timeit import default_timer as timer
import datetime

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

import re


class MetaData:

    @staticmethod
    def get_extraction_info(wait, driver, timer_start, start_time, _exit):

        try:
            wait.until(EC.presence_of_all_elements_located((By.XPATH, "//div[@class='results-summary__wrapper js-wrapper']")))
            search_result_elements = driver.find_element(By.XPATH, "//div[@class='results-summary__wrapper js-wrapper']")
            property_type = search_result_elements.find_elements(By.TAG_NAME, "span")
            property_type = str(property_type[len(property_type)-1].text)
            if len(property_type) < 1:
                property_type = "Imóvel"
        except (TimeoutException, StaleElementReferenceException, NoSuchElementException, IndexError):
            property_type = "Imóvel"

        date = datetime.date.today()

        date_time_hour = datetime.datetime.now().hour
        date_time_minute = datetime.datetime.now().minute
        date_time_second = datetime.datetime.now().second

        filename_datetime = f"{date} - {date_time_hour}h {date_time_minute}m {date_time_second}s"

        timer_end = timer()
        elapsed_seconds = timer_end - timer_start
        elapsed_time = str(datetime.timedelta(seconds=elapsed_seconds))

        end_time_hour = datetime.datetime.now().hour
        end_time_minute = datetime.datetime.now().minute
        end_time_second = datetime.datetime.now().second
        end_time = f"{end_time_hour}h {end_time_minute}m {end_time_second}s"

        try:
            wb = load_workbook(filename='Data_Extraction_Temp.xlsx')
            ws = wb.active

            # Contagem número páginas extraídas
            last_empty_row = len(list(ws.rows))
            pages = 0
            for i in range(1, last_empty_row):
                iter_ = ws[f"A{i}"].value
                match = iter_.find("Página", 0)
                if match != -1:
                    pages += 1

            found_ads = driver.find_element(By.XPATH, "//strong[@class='results-summary__count js-total-records']").text
            extracted_ads = last_empty_row - pages - 1

            # Ajustando a largura das colunas
            ws.column_dimensions['A'].width = 80
            ws.column_dimensions['B'].width = 12.5
            ws.column_dimensions['C'].width = 12.40
            ws.column_dimensions['D'].width = 14
            ws.column_dimensions['E'].width = 10.1
            ws.column_dimensions['F'].width = 50
            ws.column_dimensions['G'].width = 38.65
            ws.column_dimensions['H'].width = 133

            # Página com as informações da extração
            ws2 = wb.create_sheet('Informações da Extração')

            ws2.append(["Informações da Extração"])
            ws2.append(["Data", "Hora Início", "Hora Término", "Tempo Gasto", "Tipo de Imóvel",
                       "Quantidade de Anúncios Extraídos", "Quantidade de Páginas Extraídas"])

            #  Formatando os headers
            ws2['A1'].font = Font(bold=True, size=14)
            ws2['A1'].alignment = Alignment(horizontal='center')

            ws2['A2'].font = Font(bold=True)
            ws2['B2'].font = Font(bold=True)
            ws2['C2'].font = Font(bold=True)
            ws2['D2'].font = Font(bold=True)
            ws2['E2'].font = Font(bold=True)
            ws2['F2'].font = Font(bold=True)
            ws2['G2'].font = Font(bold=True)

            ws2.column_dimensions['A'].width = 25
            ws2.column_dimensions['B'].width = 25
            ws2.column_dimensions['C'].width = 25
            ws2.column_dimensions['D'].width = 25
            ws2.column_dimensions['E'].width = 25
            ws2.column_dimensions['F'].width = 40
            ws2.column_dimensions['G'].width = 40

            ws2.append([date, start_time, end_time, elapsed_time, property_type, f"{extracted_ads} de {found_ads} encontrados", f"{pages} páginas"])

            # Seção de validação da extração
            ws2.append(["Validação"])
            ws2.merge_cells('A1:M1')

            ws2.append([cell.value for cell in ws[3]])
            ws2.append([cell.value for cell in ws[ws.max_row // 2]])
            ws2.append([cell.value for cell in ws[ws.max_row]])

            ws2.merge_cells('A5:C5')
            ws2.merge_cells('H5:W5')
            ws2.merge_cells('A6:C6')
            ws2.merge_cells('H6:W6')
            ws2.merge_cells('A7:C7')
            ws2.merge_cells('H7:W7')

            ws2.merge_cells('A4:M4')
            ws2['A4'].font = Font(bold=True)
            ws2['A4'].alignment = Alignment(horizontal='center')

            try:
                wb.save(f".\\Extrações Viva Real\\{property_type} {filename_datetime}.xlsx")
            except FileNotFoundError:
                try:
                    property_type = property_type.replace("/", "-")
                    wb.save(f".\\{property_type} {filename_datetime}.xlsx")
                except FileNotFoundError:
                    wb.save(f".\\Extrações Viva Real\\{filename_datetime}.xlsx")

        except (FileNotFoundError, NoSuchElementException):
            if _exit.isSet():
                return
            elif not _exit.isSet():
                print("Você está tentando extrair a página errada")


class Extraction:
    """Todas as funções dessa classe tem como o objetivo extrair e/ou manipular os dados"""

    @staticmethod
    def delete_temp_sheet():
        """Deleta a planilha temporária com os dados extraídos"""
        current_dir = os.getcwd()  # Obtém o path absoluto do diretório
        cd = f'cmd /c "cd {current_dir}"'
        os.system(cd)
        os.system(r'cmd /c "del Data_Extraction_Temp.xlsx"')
        os.system('cls')

    @staticmethod
    def all_page_ads_extraction(wait, driver, _exit, pause):
        """Extrai todos os anúncios de uma página. Além disso, essa função filtra anúncios recomendados"""

        properties_addresses = []
        properties_areas = []
        properties_rooms = []
        properties_bathrooms = []
        properties_garages = []
        properties_descriptions = []
        properties_prices = []
        properties_urls = []
        properties_data = []

        try:
            os.system('cls')
            sleep(0.3)
            print("Extraindo anúncios da página...")

            #  Localizar todos os imóveis da página
            sleep(2)  # Aguardar um curto período para evitar problemas de renderização
            wait.until(EC.presence_of_all_elements_located((By.XPATH, '//div[@class="results-main__panel js-list"]')))
            properties_div = driver.find_element(By.XPATH, "//section[@class='results__main']")
            wait.until(EC.presence_of_all_elements_located((By.XPATH, "//div[@class='property-card__content']")))
            properties_data = properties_div.find_elements(By.XPATH, "//div[@class='property-card__content']")

            current_url = driver.current_url
            current_page = re.findall(r'%s(\d+)' % 'pagina=', current_url)
            current_page = f'Página {current_page}'
            if current_page == "Página []":
                current_page = "Página 1"

            # Extração dos atributos de cada imóvel
            try:
                if _exit.isSet():
                    sys.exit()

                if pause.isSet():
                    return

                # Extraindo os atributos de cada imóvel
                soup = BeautifulSoup(driver.page_source, "lxml")

                # Endereço do imóvel
                property_address_elements = soup.find_all("span", class_="property-card__address")
                for i in range(0, 36):
                    try:
                        properties_addresses.append(property_address_elements[i].get_text().strip())
                    except (IndexError, TypeError):
                        pass

                # Metragem dos imóveis
                property_area_elements = soup.find_all("li", class_="property-card__detail-item property-card__detail-area")
                for i in range(0, 36):
                    try:
                        properties_areas.append(property_area_elements[i].get_text().strip())
                    except (IndexError, TypeError):
                        pass

                # Quantidade de quartos
                property_rooms_elements = soup.find_all("li", class_="property-card__detail-item property-card__detail-room js-property-detail-rooms")
                for i in range(0, 36):
                    try:
                        properties_rooms.append(property_rooms_elements[i].get_text().strip())
                    except (IndexError, TypeError):
                        pass

                # Quantidade de banheiros
                property_bathrooms_elements = soup.find_all("li", class_="property-card__detail-item property-card__detail-bathroom js-property-detail-bathroom")
                for i in range(0, 36):
                    try:
                        properties_bathrooms.append(property_bathrooms_elements[i].get_text().strip())
                    except (IndexError, TypeError):
                        pass

                # Quantidade de vagas
                property_garage_elements = soup.find_all("li", class_="property-card__detail-item property-card__detail-garage js-property-detail-garages")
                for i in range(0, 36):
                    try:
                        properties_garages.append(property_garage_elements[i].get_text().strip())
                    except (IndexError, TypeError):
                        pass

                # Descrição do imóvel
                property_description_elements = soup.find_all("span", class_="property-card__title js-cardLink js-card-title")
                for i in range(0, 36):
                    try:
                        properties_descriptions.append(property_description_elements[i].get_text().strip())
                    except (IndexError, TypeError):
                        pass

                # Preço do imóvel
                property_price_elements = soup.find_all("div", class_="property-card__price js-property-card-prices js-property-card__price-small")
                for i in range(0, 36):
                    try:
                        properties_prices.append(property_price_elements[i].get_text().strip())
                    except (IndexError, TypeError):
                        pass

                # Link de acesso ao anúncio do imóvel
                property_url_elements = soup.find_all("a", class_="property-card__labels-container js-main-info js-listing-labels-link")
                for i in range(0, 36):
                    try:
                        properties_urls.append(f"vivareal.com.br{property_url_elements[i]['href'].strip()}")
                    except (IndexError, TypeError):
                        pass

                # Armazenando os atributos do imóvel na lista houses_data
                properties_data = list(zip(properties_addresses, properties_areas, properties_rooms, properties_bathrooms, properties_garages,
                                       properties_descriptions, properties_prices, properties_urls))

            except (TimeoutException, StaleElementReferenceException):
                # Esperar até que os elementos estejam presentes novamente
                print("Elemento tornou-se stale ou a espera excedeu o limite. Tentando próximo elemento.")
                wait.until(EC.presence_of_all_elements_located((By.XPATH, "//section[@class='results__main']")))
                wait.until(EC.presence_of_all_elements_located((By.XPATH, "//div[@class='property-card__content']")))
                sleep(2)  # Aguardar um curto período para evitar problemas de renderização

            # Salvando os dados na planilha temp
            finally:
                try:
                    wb = load_workbook('Data_Extraction_Temp.xlsx')
                    ws = wb.active
                    ws.append([current_page])
                    last_empty_row = len(list(ws.rows))
                    ws.cell(row=last_empty_row, column=1).font = Font(size=12, bold=True)

                    for row in properties_data:
                        try:
                            ws.append(row)
                        except TypeError:
                            pass

                    wb.save('Data_Extraction_Temp.xlsx')

                except FileNotFoundError:

                    wb = Workbook()
                    ws = wb.active

                    # Cabeçalho atributos dos imóveis
                    ws.cell(row=1, column=1).font = Font(bold=True)
                    ws.cell(row=1, column=1).value = "Endereço do Imóvel"

                    ws.cell(row=1, column=2).font = Font(bold=True)
                    ws.cell(row=1, column=2).value = "Área em m2"

                    ws.cell(row=1, column=3).font = Font(bold=True)
                    ws.cell(row=1, column=3).value = "Quartos"

                    ws.cell(row=1, column=4).font = Font(bold=True)
                    ws.cell(row=1, column=4).value = "Banheiros"

                    ws.cell(row=1, column=5).font = Font(bold=True)
                    ws.cell(row=1, column=5).value = "Vagas"

                    ws.cell(row=1, column=6).font = Font(bold=True)
                    ws.cell(row=1, column=6).value = "Descrição do Imóvel"

                    ws.cell(row=1, column=7).font = Font(bold=True)
                    ws.cell(row=1, column=7).value = "Preço do Imóvel(R$)"

                    ws.cell(row=1, column=8).font = Font(bold=True)
                    ws.cell(row=1, column=8).value = "Link Anúncio"

                    ws.append([current_page])
                    last_empty_row = len(list(ws.rows))
                    ws.cell(row=last_empty_row, column=1).font = Font(bold=True)

                    for row in properties_data:
                        try:
                            ws.append(row)
                        except TypeError:
                            pass

                    wb.save('Data_Extraction_Temp.xlsx')

                sleep(1.5)
                #os.system('cls')
                print("FIM extração")
                print("Próxima página liberada")
        except (urllib3.exceptions.MaxRetryError, urllib3.exceptions.ProtocolError, ConnectionResetError):
            print("Extração Interrompida")
        finally:
            return len(properties_data)













